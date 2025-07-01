import tkinter as tk
from tkinter import ttk, messagebox, TclError, simpledialog
from tkinter.constants import ANCHOR, END
import pyupbit
import pandas as pd
import numpy as np
import mplfinance as mpf
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import threading
import time
import platform
import os
import requests
from queue import Queue, Empty
import json
from datetime import datetime, timedelta
from scipy.signal import find_peaks
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import filedialog   # ← 추가

# ----------------------------------------------------------------------------- 
# 한글 폰트 설정
# -----------------------------------------------------------------------------
if platform.system() == 'Windows':
    plt.rc('font', family='Malgun Gothic')
elif platform.system() == 'Darwin':  # Mac OS
    plt.rc('font', family='AppleGothic')
else:  # Linux
    plt.rc('font', family='NanumGothic')
plt.rcParams['axes.unicode_minus'] = False

# ----------------------------------------------------------------------------- 
# 1. API Key와 기본 설정
# ----------------------------------------------------------------------------- 
def select_login_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="login.txt 파일을 선택하세요",
        filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")]
    )
    root.destroy()
    return file_path

login_file = select_login_file()
if not login_file:
    messagebox.showerror("로그인 파일 오류", "login.txt 파일을 선택하지 않았습니다. 프로그램을 종료합니다.")
    exit()

try:
    with open(login_file, "r") as f:
        lines = f.readlines()
        if len(lines) < 3:
            raise ValueError("파일에 access key, secret key, 자동매매 비밀번호가 모두 필요합니다.")
        access = lines[0].strip()
        secret = lines[1].strip()
        trade_password = lines[2].strip() 
except FileNotFoundError:
    messagebox.showerror("로그인 파일 오류", "login.txt 파일을 찾을 수 없습니다.\n선택한 파일을 확인해주세요.")
    exit()
except Exception as e:
    messagebox.showerror("로그인 파일 오류", f"login.txt 파일 처리 중 오류가 발생했습니다.\n\n{e}")
    exit()

try:
    upbit = pyupbit.Upbit(access, secret)
    balances = upbit.get_balances()
    print("✅ 업비트 로그인 성공")
except Exception as e:
    messagebox.showerror("로그인 실패", f"API 키가 유효하지 않거나 네트워크에 문제가 있습니다.\nlogin.txt 파일을 확인해주세요.\n\n{e}")
    exit()

# ----------------------------------------------------------------------------- 
# 2. GUI 클래스 및 기능
# -----------------------------------------------------------------------------
class UpbitChartApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.trade_password = trade_password
        self.auto_trade_monitor = None  # 반드시 선언!
        self.title("업비트 포트폴리오 & HTS")
        self.geometry("1600x980")
        self.selected_ticker_display = tk.StringVar()
        self.selected_interval = tk.StringVar(value='day')
        self.current_price = 0.0; self.avg_buy_price = 0.0
        self.balances_data = {}; self.is_running = True
        self.chart_elements = {'main': [], 'overlay': []}
        self.krw_balance_summary_var = tk.StringVar(value="보유 KRW: 0 원")
        self.total_investment_var = tk.StringVar(value="총 투자금액: 0 원")
        self.total_valuation_var = tk.StringVar(value="총 평가금액: 0 원")
        self.total_pl_var = tk.StringVar(value="총 평가손익: 0 원 (0.00%)")
        self.is_panning = False; self.pan_start_pos = None
        self.master_df = None
        self.is_loading_older = False
        self.data_bounds = {'x': None, 'y': None}
        self.ma_vars = {'5': tk.BooleanVar(value=True), '20': tk.BooleanVar(value=True), '60': tk.BooleanVar(), '120': tk.BooleanVar()}
        self.bb_var = tk.BooleanVar(value=True)
        self.ticker_to_display_name = {}
        self.display_name_to_ticker = {}
        self.load_ticker_names()
        self.market_data = []
        self.sort_column = 'volume'; self.sort_ascending = False
        self.buy_order_type = tk.StringVar(value="limit")
        self.buy_price_var = tk.StringVar(); self.buy_amount_var = tk.StringVar(); self.buy_total_var = tk.StringVar()
        self.sell_order_type = tk.StringVar(value="limit")
        self.sell_price_var = tk.StringVar(); self.sell_amount_var = tk.StringVar(); self.sell_total_var = tk.StringVar()
        self.sell_percentage_var = tk.StringVar()
        self.krw_balance = 0.0
        self.coin_balance = 0.0
        self.buy_krw_balance_var = tk.StringVar(value="주문가능: 0 KRW")
        self.sell_coin_balance_var = tk.StringVar(value="주문가능: 0 COIN")
        self._is_calculating = False
        self.data_queue = Queue()
        self.trade_history_data = []
        self.is_auto_trading = False
        self.auto_trade_settings = {}
        self.auto_trade_thread = None
        self.last_sell_time = {}
        self.load_auto_trade_settings()
        self.create_widgets()
        self.add_variable_traces()
        self.load_trade_history_from_file()
        self.update_history_tree_gui()
        self.load_my_tickers()
        self.update_loop_counter = 0
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self._keep_view = False  # 차트 뷰 상태 유지용 플래그
        self._ignore_market_select_event = False  # 마켓 트리 이벤트 루프 방지

    def load_auto_trade_settings(self):
        try:
            with open("auto_trade_settings.json", "r", encoding="utf-8") as f:
                self.auto_trade_settings = json.load(f)
                print("✅ 자동매매 설정 로드 완료.")
        except (FileNotFoundError, json.JSONDecodeError):
            print("ℹ️ 자동매매 설정 파일 없음. 기본값으로 시작합니다.")
            self.auto_trade_settings = {
                'enabled_tickers': [],
                'investment_ratio': 10, 
                'is_unowned_buy_enabled': False,
                'cooldown_hours': 24,

                'buy_scenarios': {
                    'consolidation': {'enabled': True},
                    'reversal': {'enabled': False}
                },
                'sell_scenarios': {
                    'overheat': {'enabled': True},
                    'divergence': {'enabled': True}
                }
            }
            self.save_auto_trade_settings()

    def save_auto_trade_settings(self):
        with open("auto_trade_settings.json", "w", encoding="utf-8") as f:
            json.dump(self.auto_trade_settings, f, ensure_ascii=False, indent=4)
        print("💾 자동매매 설정 저장 완료.")

    def start_updates(self):
        self.update_loop()
        self.process_queue()

    def load_ticker_names(self):
        print("🔍 종목 이름 정보를 로드합니다...")
        try:
            url = "https://api.upbit.com/v1/market/all?isDetails=true"
            response = requests.get(url)
            response.raise_for_status()
            all_market_info = response.json()
            for market_info in all_market_info:
                if market_info['market'].startswith('KRW-'):
                    market = market_info['market']
                    korean_name = market_info['korean_name']
                    symbol = market.split('-')[1]
                    display_name = f"{korean_name}({symbol})"
                    self.ticker_to_display_name[market] = display_name
                    self.display_name_to_ticker[display_name] = market
            print("✅ 종목 이름 정보 로드 완료.")
        except Exception as e:
            print(f"❗️ 종목 이름 정보 로드 실패: {e}\n종목 코드를 그대로 사용합니다.")
            
    def add_variable_traces(self):
        self.buy_price_var.trace_add("write", self._on_buy_input_change)
        self.buy_amount_var.trace_add("write", self._on_buy_input_change)
        self.buy_total_var.trace_add("write", self._on_buy_total_change)
        self.sell_price_var.trace_add("write", self._on_sell_input_change)
        self.sell_amount_var.trace_add("write", self._on_sell_input_change)
        
    def create_widgets(self):
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(main_frame, width=500, padding=10)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        top_left_frame = ttk.Frame(left_frame)
        top_left_frame.pack(fill=tk.X, pady=5)

        summary_frame = ttk.LabelFrame(top_left_frame, text="종합 현황", padding=10)
        summary_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        ttk.Label(summary_frame, textvariable=self.krw_balance_summary_var, font=("Helvetica", 12)).pack(anchor="w")
        ttk.Label(summary_frame, textvariable=self.total_investment_var, font=("Helvetica", 12)).pack(anchor="w")
        ttk.Label(summary_frame, textvariable=self.total_valuation_var, font=("Helvetica", 12)).pack(anchor="w")
        self.total_pl_label = ttk.Label(summary_frame, textvariable=self.total_pl_var, font=("Helvetica", 12, "bold"))
        self.total_pl_label.pack(anchor="w")

        pie_frame = ttk.LabelFrame(top_left_frame, text="코인 비중", padding=10)
        pie_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        self.pie_fig, self.pie_ax = plt.subplots(figsize=(3, 2.5))
        self.pie_fig.patch.set_facecolor('#F0F0F0')
        self.pie_canvas = FigureCanvasTkAgg(self.pie_fig, master=pie_frame)
        self.pie_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        tree_frame = ttk.LabelFrame(left_frame, text="보유 코인 (더블클릭하여 차트 보기)", padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        cols = ('display_name', 'balance', 'avg_price', 'cur_price', 'valuation', 'pl', 'pl_rate')
        self.portfolio_tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        col_map = {"종목명": 120, "보유수량": 80, "매수평균가": 90, "현재가": 90, "평가금액": 90, "평가손익": 90, "손익(%)": 70}
        for i, (text, width) in enumerate(col_map.items()):
            self.portfolio_tree.heading(cols[i], text=text)
            self.portfolio_tree.column(cols[i], width=width, anchor='e')
        self.portfolio_tree.column('display_name', anchor='w')
        
        self.portfolio_tree.tag_configure('plus', foreground='red')
        self.portfolio_tree.tag_configure('minus', foreground='blue')
        
        self.portfolio_tree.pack(fill=tk.BOTH, expand=True)
        self.portfolio_tree.bind("<Double-1>", self.on_tree_double_click)

        # --- [수정] 로그 프레임을 좌측 하단이 아닌, 좌측 프레임의 맨 아래에 높이만 고정해서 배치 ---
        log_frame = ttk.LabelFrame(left_frame, text="자동매매 로그", padding=10)
        log_frame.pack(side=tk.BOTTOM, fill=tk.X, expand=False, pady=(5,0))  # fill=tk.X, expand=False로 변경
        log_text_frame = ttk.Frame(log_frame)
        log_text_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_text_frame, height=7, state='disabled', font=('Courier New', 10), wrap='none')  # height를 7~8로 조정
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scrollbar = ttk.Scrollbar(log_text_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        log_xscrollbar = ttk.Scrollbar(log_text_frame, orient="horizontal", command=self.log_text.xview)
        self.log_text.configure(xscrollcommand=log_xscrollbar.set)
        log_xscrollbar.pack(side=tk.BOTTOM, fill=tk.X)

        # [추가] 자동매매 모니터링 창 다시 띄우기 버튼
        monitor_btn_frame = ttk.Frame(log_frame)
        monitor_btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(monitor_btn_frame, text="모니터링 창 열기", command=self.show_auto_trade_monitor).pack(side=tk.RIGHT, padx=5)

        # --- 이하 기존 코드(오더, 히스토리, 우측 프레임 등) 동일 ---
        bottom_frame = ttk.Frame(left_frame)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, pady=(5,0))
        
        bottom_left_frame = ttk.Frame(bottom_frame)
        bottom_left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        bottom_controls_frame = ttk.Frame(bottom_left_frame)
        bottom_controls_frame.pack(side=tk.TOP, fill=tk.X)

        self.order_notebook = ttk.Notebook(bottom_controls_frame)
        self.order_notebook.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, ipady=5)

        buy_tab = ttk.Frame(self.order_notebook, padding=10)
        sell_tab = ttk.Frame(self.order_notebook, padding=10)
        history_tab = ttk.Frame(self.order_notebook, padding=10)
        self.order_notebook.add(buy_tab, text="매수")
        self.order_notebook.add(sell_tab, text="매도")
        self.order_notebook.add(history_tab, text="거래내역")
        self.create_buy_sell_tab(buy_tab, "buy")
        self.create_buy_sell_tab(sell_tab, "sell")
        self.create_history_tab(history_tab)
        
        auto_trade_frame = ttk.LabelFrame(bottom_controls_frame, text="자동매매", padding=10)
        auto_trade_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5,0))
        
        ttk.Style().configure("On.TButton", foreground="black", background="#4CAF50", font=('Helvetica', 10, 'bold'))
        ttk.Style().configure("Off.TButton", foreground="black", background="#F44336", font=('Helvetica', 10, 'bold'))

        self.auto_trade_toggle_button = ttk.Button(auto_trade_frame, text="자동매매 켜기", style="Off.TButton", command=self.toggle_auto_trading)
        self.auto_trade_toggle_button.pack(pady=10, padx=10, fill='x', ipady=8)

        settings_button = ttk.Button(auto_trade_frame, text="자동매매 설정", command=self.open_settings_window)
        settings_button.pack(pady=5, padx=10, fill='x')
        
        # log_frame = ttk.LabelFrame(bottom_left_frame, text="자동매매 로그", padding=10)
        # log_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, pady=(5,0))
        # self.log_text = tk.Text(log_frame, height=5, state='disabled', font=('Courier New', 9))
        # self.log_text.pack(fill=tk.BOTH, expand=True)

        right_frame = ttk.Frame(main_frame, padding=10)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        market_list_frame = ttk.LabelFrame(right_frame, text="KRW 마켓 (컬럼 헤더 클릭하여 정렬)", padding=10)
        market_list_frame.pack(side="top", fill="x", pady=5)
        market_cols = ('display_name', 'price', 'change_rate', 'volume')
        self.market_tree = ttk.Treeview(market_list_frame, columns=market_cols, show='headings', height=5)
        market_col_map = {"종목명": 150, "현재가": 100, "등락률": 80, "거래대금(24h)": 120}
        for i, (text, width) in enumerate(market_col_map.items()):
            col_id = market_cols[i]
            self.market_tree.heading(col_id, text=text, command=lambda c=col_id: self.sort_market_list(c))
            self.market_tree.column(col_id, width=width, anchor='e')
        self.market_tree.column('display_name', anchor='w')
        scrollbar = ttk.Scrollbar(market_list_frame, orient="vertical", command=self.market_tree.yview)
        self.market_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y"); self.market_tree.pack(side="left", fill="both", expand=True)
        self.market_tree.tag_configure('red', foreground='red'); self.market_tree.tag_configure('blue', foreground='blue'); self.market_tree.tag_configure('black', foreground='black')
        self.market_tree.bind("<<TreeviewSelect>>", self.on_market_list_select)
        
        control_frame_1 = ttk.Frame(right_frame); control_frame_1.pack(side="top", fill="x", pady=(10,0))
        control_frame_2 = ttk.Frame(right_frame); control_frame_2.pack(side="top", fill="x", pady=5)
        ttk.Label(control_frame_1, text="종목 선택:").pack(side="left")
        self.ticker_combobox = ttk.Combobox(control_frame_1, textvariable=self.selected_ticker_display, width=20)
        self.ticker_combobox.pack(side="left", padx=(5, 15)); self.ticker_combobox.bind("<<ComboboxSelected>>", self.on_ticker_select)
        ttk.Label(control_frame_1, text="차트 주기:").pack(side="left")
        intervals = {"5분봉": "minute5", "30분봉": "minute30", "1시간봉": "minute60", "4시간봉": "minute240", "일봉": "day", "주봉": "week"}
        for text, value in intervals.items():
            rb = ttk.Radiobutton(control_frame_1, text=text, variable=self.selected_interval, value=value, command=self.on_ticker_select)
            rb.pack(side="left")
        ttk.Label(control_frame_2, text="보조지표: ").pack(side="left")
        for period, var in self.ma_vars.items():
            cb = ttk.Checkbutton(control_frame_2, text=f"MA{period}", variable=var, command=self.draw_base_chart)
            cb.pack(side="left")
        cb_bb = ttk.Checkbutton(control_frame_2, text="BBands", variable=self.bb_var, command=self.draw_base_chart)
        cb_bb.pack(side="left", padx=5)

        chart_frame = ttk.Frame(right_frame); chart_frame.pack(side="bottom", fill="both", expand=True, pady=5)
        self.fig, self.ax = plt.subplots(figsize=(10, 6)); self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.canvas.mpl_connect('scroll_event', self.on_scroll); self.canvas.mpl_connect('button_press_event', self.on_press)
        self.canvas.mpl_connect('motion_notify_event', self.on_motion); self.canvas.mpl_connect('button_release_event', self.on_release)

    def show_auto_trade_monitor(self):
        if self.auto_trade_monitor is None or not self.auto_trade_monitor.winfo_exists():
            self.auto_trade_monitor = AutoTradeMonitorWindow(self)
        else:
            self.auto_trade_monitor.lift()

    def log_auto_trade(self, message, log_dict=None):
        now = datetime.now().strftime('%H:%M:%S')
        log_message = f"[{now}] {message}"
        
        self.log_text.config(state='normal')
        self.log_text.insert(END, log_message + "\n")
        self.log_text.see(END)
        self.log_text.config(state='disabled')
        
        print(log_message)
        # --- 자동매매 모니터에도 기록 ---
        if log_dict and self.auto_trade_monitor is not None and self.auto_trade_monitor.winfo_exists():
            self.auto_trade_monitor.add_log(log_dict)
    
    def toggle_auto_trading(self):
        if not self.is_auto_trading:
            entered_password = simpledialog.askstring("비밀번호 확인", "자동매매를 시작하려면 비밀번호 4자리를 입력하세요.", show='*')
            if entered_password is None: return
            if entered_password == self.trade_password:
                self.is_auto_trading = True
                self.auto_trade_toggle_button.config(text="자동매매 끄기", style="On.TButton")
                self.log_auto_trade("▶️ 자동매매 시작")
                # --- [수정] 자동매매 옵션 정보 로그 출력 ---
                s = self.auto_trade_settings
                enabled_tickers = s.get('enabled_tickers', [])
                buy_ratio = s.get('investment_ratio', 10)
                cooldown = s.get('cooldown_hours', 24)
                unowned = s.get('is_unowned_buy_enabled', False)
                # 전략 선택 정보 추가
                strategy_names = []
                if s.get('strategy1'):
                    strategy_names.append("전략1(RSI+이동평균+거래량)")
                if s.get('strategy2'):
                    strategy_names.append("전략2(볼린저밴드+캔들패턴)")
                if s.get('strategy3'):
                    strategy_names.append("전략3(MACD+트레일링스탑)")
                if s.get('strategy4'):
                    strategy_names.append("전략4(강한캔들+볼륨펌핑)")
                strategy_desc = ', '.join(strategy_names) if strategy_names else '없음'

                buy_scenarios = []
                if s.get('buy_scenarios', {}).get('consolidation', {}).get('enabled', False):
                    buy_scenarios.append("상승추세 눌림목")
                if s.get('buy_scenarios', {}).get('reversal', {}).get('enabled', False):
                    buy_scenarios.append("바닥권 추세전환")
                sell_scenarios = []
                if s.get('sell_scenarios', {}).get('overheat', {}).get('enabled', False):
                    sell_scenarios.append("과열 고점")
                if s.get('sell_scenarios', {}).get('divergence', {}).get('enabled', False):
                    sell_scenarios.append("하락 다이버전스")
                if unowned:
                    target_desc = "미보유 KRW마켓 전체"
                else:
                    target_desc = ', '.join([self.ticker_to_display_name.get(t, t) for t in enabled_tickers]) if enabled_tickers else '없음'

                self.log_auto_trade(
                    f"설정 요약: "
                    f"매수비중 {buy_ratio}%, 쿨다운 {cooldown}시간, "
                    f"미보유코인매수 {'ON' if unowned else 'OFF'}, "
                    f"전략: {strategy_desc}, "
                    f"매수시나리오: {', '.join(buy_scenarios) if buy_scenarios else '없음'}, "
                    f"매도시나리오: {', '.join(sell_scenarios) if sell_scenarios else '없음'}, "
                    f"대상종목: {target_desc}"
                )
                self.show_auto_trade_monitor()  # 자동매매 시작 시 모니터 창 자동 표시
                self.auto_trade_thread = threading.Thread(target=self.auto_trade_worker, daemon=True)
                self.auto_trade_thread.start()
            else:
                messagebox.showerror("인증 실패", "비밀번호가 일치하지 않습니다.")
        else:
            self.is_auto_trading = False
            self.auto_trade_toggle_button.config(text="자동매매 켜기", style="Off.TButton")
            self.log_auto_trade("⏹️ 자동매매 중지")

    def open_settings_window(self):
        settings_window = AutoTradeSettingsWindow(self)
        settings_window.grab_set()

    def get_technical_indicators(self, ticker, interval='day', count=200):
        try:
            df = pyupbit.get_ohlcv(ticker, interval=interval, count=count)
            return self.get_technical_indicators_from_raw(df)
        except Exception as e:
            self.log_auto_trade(f"❗️ {ticker} 지표 계산 오류: {e}")
            return None
    
    # <<<<< [핵심 수정] 과거 데이터 로딩 및 신규 상장 코인 차트 표시 개선 >>>>>
    def _fetch_older_data_worker(self, ticker, interval, to_date, current_xlim):
        try:
            # pyupbit의 to 파라미터는 해당 시점 '이전'까지 200개를 반환하므로, to_date에서 1초 빼기
            if isinstance(to_date, pd.Timestamp):
                to_date_str = (to_date - pd.Timedelta(seconds=1)).strftime('%Y-%m-%d %H:%M:%S')
            else:
                to_date_str = pd.to_datetime(to_date) - pd.Timedelta(seconds=1)
                to_date_str = to_date_str.strftime('%Y-%m-%d %H:%M:%S')

            # 과거 데이터 요청 (최대 200개)
            older_df_raw = pyupbit.get_ohlcv(ticker, interval=interval, count=200, to=to_date_str)
            # 만약 데이터가 없거나 1개 이하만 반환되면, 더 이상 로드할 데이터가 없음
            if older_df_raw is None or len(older_df_raw) < 2:
                print("ℹ️ 더 이상 로드할 과거 데이터가 없습니다.")
                self.is_loading_older = False
                return

            # 현재 데이터에서 순수 OHLCV만 추출
            current_ohlcv = self.master_df[['open', 'high', 'low', 'close', 'volume']]
            # 중복 인덱스 제거 (과거 데이터와 현재 데이터가 겹칠 수 있음)
            combined_df_raw = pd.concat([older_df_raw, current_ohlcv])
            combined_df_raw = combined_df_raw[~combined_df_raw.index.duplicated(keep='last')]
            # 인덱스 정렬
            combined_df_raw = combined_df_raw.sort_index()

            # 지표 계산 최소 데이터 개수 완화 (신규상장 코인 대응)
            df_with_indicators = self.get_technical_indicators_from_raw(combined_df_raw, min_length=2)
            # 기존 master_df보다 더 많은 데이터가 있으면 갱신
            if df_with_indicators is not None and not df_with_indicators.empty:
                num_candles_added = len(df_with_indicators) - len(self.master_df)
                if num_candles_added > 0:
                    # xlim을 왼쪽으로 이동
                    new_xlim = (current_xlim[0] + num_candles_added, current_xlim[1] + num_candles_added)
                    self.data_queue.put(("draw_older_chart", (df_with_indicators, new_xlim)))
                    return
            print("ℹ️ 더 이상 로드할 과거 데이터가 없습니다.")
        except Exception as e:
            print(f"❗️ 과거 데이터 로딩 중 오류 발생: {e}")
        self.is_loading_older = False

    # get_technical_indicators_from_raw의 최소 데이터 개수 파라미터화 및 신규상장 코인 대응
    def get_technical_indicators_from_raw(self, df, min_length=2):
        if df is None or len(df) < min_length:
            return None
        for p in [5, 20, 60, 120]:
            df[f'ma{p}'] = df['close'].rolling(window=p, min_periods=1).mean()
        delta = df['close'].diff(1)
        gain = (delta.where(delta > 0, 0)).rolling(window=14, min_periods=1).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=14, min_periods=1).mean()
        rs = gain / loss.replace(0, np.nan)
        df['rsi'] = 100 - (100 / (1 + rs))
        df['ema12'] = df['close'].ewm(span=12, adjust=False, min_periods=1).mean()
        df['ema26'] = df['close'].ewm(span=26, adjust=False, min_periods=1).mean()
        df['macd'] = df['ema12'] - df['ema26']
        df['signal'] = df['macd'].ewm(span=9, adjust=False, min_periods=1).mean()
        df['body'] = abs(df['close'] - df['open'])
        df['upper_shadow'] = df['high'] - df[['open', 'close']].max(axis=1)
        df['lower_shadow'] = df[['open', 'close']].min(axis=1) - df['low']
        df['is_green'] = df['close'] > df['open']
        patterns = []
        for i in range(1, len(df)):
            prev, curr = df.iloc[i-1], df.iloc[i]
            pattern = 'none'
            if curr['is_green'] and (curr['lower_shadow'] > curr['body'] * 2) and (curr['upper_shadow'] < curr['body'] * 0.5): pattern = 'hammer'
            elif not prev['is_green'] and curr['is_green'] and curr['close'] > prev['open'] and curr['open'] < prev['close']: pattern = 'bullish_engulfing'
            elif not curr['is_green'] and (curr['upper_shadow'] > curr['body'] * 2) and (curr['lower_shadow'] < curr['body'] * 0.5): pattern = 'shooting_star'
            patterns.append(pattern)
        df['pattern'] = ['none'] + patterns if len(df) > 1 else ['none']
        try:
            rsi_peaks, _ = find_peaks(df['rsi'].fillna(0), distance=5, width=1)
            rsi_troughs, _ = find_peaks(-df['rsi'].fillna(0), distance=5, width=1)
        except Exception:
            rsi_peaks, rsi_troughs = [], []
        # --- 핵심 수정: _check_divergence를 인스턴스 메서드가 아닌 staticmethod로 분리 ---
        df['bearish_div'] = UpbitChartApp._check_divergence_static(df, rsi_peaks, 'bearish')
        df['bullish_div'] = UpbitChartApp._check_divergence_static(df, rsi_troughs, 'bullish')
        return df

    @staticmethod
    def _check_divergence_static(df, peaks, div_type):
        # rsi와 가격의 피크/트로프를 비교하여 다이버전스 여부를 반환
        if len(peaks) < 2:
            return [False] * len(df)
        result = [False] * len(df)
        for i in range(1, len(peaks)):
            idx1, idx2 = peaks[i-1], peaks[i]
            if div_type == 'bearish':
                # RSI는 하락, 가격은 상승
                if df['rsi'].iloc[idx2] < df['rsi'].iloc[idx1] and df['high'].iloc[idx2] > df['high'].iloc[idx1]:
                    result[idx2] = True
            elif div_type == 'bullish':
                # RSI는 상승, 가격은 하락
                if df['rsi'].iloc[idx2] > df['rsi'].iloc[idx1] and df['low'].iloc[idx2] < df['low'].iloc[idx1]:
                    result[idx2] = True
        return result

    def auto_trade_worker(self):
        self.log_auto_trade(f"자동매매 스레드 시작.")
        while self.is_auto_trading:
            try:
                s = self.auto_trade_settings
                enabled_tickers = s.get('enabled_tickers', [])
                if enabled_tickers:
                    ticker = enabled_tickers[0]
                    my_coins = self.balances_data
                    coin_info = my_coins.get(ticker)
                    if coin_info and float(coin_info.get('balance', 0)) > 0:
                        self.check_sell_condition(ticker, coin_info)
                # 전략별 분기 실행 (run_strategyX 내부에서 매수/매도 모두 처리)
                if s.get('strategy1'):
                    self.run_strategy1()
                if s.get('strategy2'):
                    self.run_strategy2()
                if s.get('strategy3'):
                    self.run_strategy3()
                if s.get('strategy4'):
                    self.run_strategy4()
                if s.get('strategy5'):
                    self.run_strategy5()
                if s.get('strategy6'):
                    self.run_strategy6()
                if s.get('strategy7'):
                    self.run_strategy7()
                if s.get('strategy8'):
                    self.run_strategy8()
            except Exception as e:
                import traceback
                self.log_auto_trade(f"❗️ 자동매매 루프 오류: {e}\n{traceback.format_exc()}")
            time.sleep(60)

    # 3. [UpbitChartApp] 각 전략별 함수 뼈대 추가 (조건 구현은 각 전략 설명 참고)
    def run_strategy1(self):
        # 5분봉, RSI+이동평균+거래량 전략 (저위험)
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략1", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute5', count=200)
        if df is None or len(df) < 30:
            return
        last = df.iloc[-1]
        prev = df.iloc[-2]

        # 매수: RSI 30 이하 & 5MA > 20MA & 거래량 최근 10봉 평균의 2배 이상
        vol_ma10 = df['volume'].rolling(10).mean().iloc[-1]
        if last['rsi'] < 30 and last['ma5'] > last['ma20'] and last['volume'] > vol_ma10 * 2:
            self.execute_buy(ticker, "전략1: RSI<30 & 5MA>20MA & 거래량급증")
            return

        # 매도: RSI 70 이상 & 5MA < 20MA
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if last['rsi'] > 70 and last['ma5'] < last['ma20']:
                self.execute_sell(ticker, coin_info, "전략1: RSI>70 & 5MA<20MA")

    def run_strategy2(self):
        # 1분봉, 볼린저밴드+캔들패턴 전략 (고수익)
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략2", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute1', count=100)
        if df is None or len(df) < 30:
            return
        last = df.iloc[-1]
        prev = df.iloc[-2]
        bb_period = 20
        middle = df['close'].rolling(window=bb_period).mean()
        std = df['close'].rolling(window=bb_period).std()
        upper = middle + (std * 2)
        lower = middle - (std * 2)

        # 매수: 종가가 볼린저밴드 하단 돌파 + 해머형 캔들
        if last['close'] < lower.iloc[-1] and last['pattern'] == 'hammer':
            self.execute_buy(ticker, "전략2: BB하단돌파+해머형")
            return

        # 매도: 종가가 볼린저밴드 상단 돌파 + shooting_star
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if last['close'] > upper.iloc[-1] and last['pattern'] == 'shooting_star':
                self.execute_sell(ticker, coin_info, "전략2: BB상단돌파+슈팅스타")

    def run_strategy3(self):
        # 5분봉, MACD+트레일링스탑 전략 (추세추종)
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략3", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute5', count=100)
        if df is None or len(df) < 30:
            return
        last = df.iloc[-1]
        prev = df.iloc[-2]

        # 매수: MACD 시그널 상향돌파
        if prev['macd'] < prev['signal'] and last['macd'] > last['signal']:
            self.execute_buy(ticker, "전략3: MACD상향돌파")
            return

        # 매도: MACD 시그널 하향돌파 or 트레일링스탑(최고가 대비 -3% 하락)
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            buy_price = float(coin_info.get('avg_buy_price', 0))
            max_high = df['high'].iloc[-20:].max()
            cur_price = last['close']
            if prev['macd'] > prev['signal'] and last['macd'] < last['signal']:
                self.execute_sell(ticker, coin_info, "전략3: MACD하향돌파")
            elif buy_price > 0 and cur_price < max_high * 0.97:
                self.execute_sell(ticker, coin_info, "전략3: 트레일링스탑 -3%")

    def run_strategy4(self):
        # 1분봉, 강한캔들+볼륨펌핑 전략 (초단타)
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략4", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute1', count=30)
        if df is None or len(df) < 10:
            return
        last = df.iloc[-1]
        prev = df.iloc[-2]
        # 매수: 전봉 대비 거래량 3배 이상 + 양봉 + 몸통 길이 상위 20%
        body_threshold = df['body'].quantile(0.8)
        if last['is_green'] and last['volume'] > prev['volume'] * 3 and last['body'] > body_threshold:
            self.execute_buy(ticker, "전략4: 강한양봉+볼륨펌핑")
            return

        # 매도: 음봉 + 거래량 급증
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if not last['is_green'] and last['volume'] > prev['volume'] * 2:
                self.execute_sell(ticker, coin_info, "전략4: 강한음봉+볼륨급증")

    def run_strategy5(self):
        # MA 골든/데드크로스 (5/20MA)
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략5", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute5', count=100)
        if df is None or len(df) < 30:
            return
        last = df.iloc[-1]
        prev = df.iloc[-2]
        # 골든크로스 매수
        if prev['ma5'] < prev['ma20'] and last['ma5'] > last['ma20']:
            self.execute_buy(ticker, "전략5: MA 골든크로스")
            return
        # 데드크로스 매도
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if prev['ma5'] > prev['ma20'] and last['ma5'] < last['ma20']:
                self.execute_sell(ticker, coin_info, "전략5: MA 데드크로스")

    def run_strategy6(self):
        # OBV 추세전환
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략6", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute5', count=100)
        if df is None or len(df) < 30:
            return
        # OBV 계산
        obv = [0]
        for i in range(1, len(df)):
            if df['close'].iloc[i] > df['close'].iloc[i-1]:
                obv.append(obv[-1] + df['volume'].iloc[i])
            elif df['close'].iloc[i] < df['close'].iloc[i-1]:
                obv.append(obv[-1] - df['volume'].iloc[i])
            else:
                obv.append(obv[-1])
        df['obv'] = obv
        last = df.iloc[-1]
        prev = df.iloc[-2]
        # 매수: OBV가 직전 저점 돌파(상승 전환)
        if df['obv'].iloc[-2] < df['obv'].iloc[-3] and df['obv'].iloc[-1] > df['obv'].iloc[-2]:
            self.execute_buy(ticker, "전략6: OBV 상승전환")
            return
        # 매도: OBV가 직전 고점 하락(하락 전환)
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if df['obv'].iloc[-2] > df['obv'].iloc[-3] and df['obv'].iloc[-1] < df['obv'].iloc[-2]:
                self.execute_sell(ticker, coin_info, "전략6: OBV 하락전환")

    def run_strategy7(self):
        # StochRSI 돌파
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략7", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute5', count=100)
        if df is None or len(df) < 30:
            return
        # StochRSI 계산
        rsi = df['rsi']
        min_rsi = rsi.rolling(window=14, min_periods=1).min()
        max_rsi = rsi.rolling(window=14, min_periods=1).max()
        stochrsi = (rsi - min_rsi) / (max_rsi - min_rsi)
        df['stochrsi'] = stochrsi
        last = df.iloc[-1]
        prev = df.iloc[-2]
        # 매수: StochRSI 0.2 이하에서 상향 돌파
        if prev['stochrsi'] < 0.2 and last['stochrsi'] >= 0.2:
            self.execute_buy(ticker, "전략7: StochRSI 상향돌파")
            return
        # 매도: StochRSI 0.8 이상에서 하향 돌파
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if prev['stochrsi'] > 0.8 and last['stochrsi'] <= 0.8:
                self.execute_sell(ticker, coin_info, "전략7: StochRSI 하향돌파")

    def run_strategy8(self):
        # CCI 돌파
        s = self.auto_trade_settings
        ticker = s.get('enabled_tickers', [None])[0]
        if not ticker:
            return
        cooldown_minutes = 10  # 전략1 쿨다운 10분
        if self.is_cooldown(ticker, "전략8", cooldown_minutes):
            return
        df = self.get_technical_indicators(ticker, interval='minute5', count=100)
        if df is None or len(df) < 30:
            return
        # CCI 계산
        tp = (df['high'] + df['low'] + df['close']) / 3
        ma = tp.rolling(window=20, min_periods=1).mean()
        md = tp.rolling(window=20, min_periods=1).apply(lambda x: np.mean(np.abs(x - np.mean(x))))
        cci = (tp - ma) / (0.015 * md)
        df['cci'] = cci
        last = df.iloc[-1]
        prev = df.iloc[-2]
        # 매수: CCI 100 돌파
        if prev['cci'] < 100 and last['cci'] >= 100:
            self.execute_buy(ticker, "전략8: CCI 100 상향돌파")
            return
        # 매도: CCI -100 돌파
        coin_info = self.balances_data.get(ticker)
        if coin_info and float(coin_info.get('balance', 0)) > 0:
            if prev['cci'] > -100 and last['cci'] <= -100:
                self.execute_sell(ticker, coin_info, "전략8: CCI -100 하향돌파")

    def execute_sell(self, ticker, coin_info, reason):
        self.log_auto_trade(f"📉 [{reason}] {ticker} 매도 신호 포착")
        balance = float(coin_info['balance'])
        price = pyupbit.get_current_price(ticker)  # 현재가 가져오기
        try:
            upbit.sell_market_order(ticker, balance)
            self.last_sell_time[ticker] = datetime.now()
            entry_price = float(coin_info['avg_buy_price'])
            total_buy = entry_price * balance
            total_sell = price * balance if price else 0
            profit = total_sell - total_buy if price else ""
            profit_rate = ((price - entry_price) / entry_price * 100) if price else ""
            # 업비트 수수료율(시장가 0.05%) 적용
            fee_buy = total_buy * 0.0005
            fee_sell = total_sell * 0.0005
            total_fee = fee_buy + fee_sell
            profit_net = profit - total_fee if price else ""
            profit_rate_net = ((profit_net / total_buy) * 100) if price else ""
            self.log_auto_trade(
                f"✅ {ticker} 전량 매도 주문 완료.",
                log_dict={
                    "datetime": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "ticker": ticker,
                    "side": "매도",
                    "reason": reason,
                    "entry_price": entry_price,
                    "amount": balance,
                    "total": total_buy,
                    "exit_price": price,
                    "profit": profit,
                    "profit_rate": profit_rate,
                    "fee": total_fee,
                    "profit_rate_net": profit_rate_net
                }
            )
        except Exception as e:
            self.log_auto_trade(f"❗️ {ticker} 매도 주문 실행 중 오류: {e}")

    def execute_buy(self, ticker, reason):
        try:
            krw_balance = upbit.get_balance("KRW")
            ratio = self.auto_trade_settings.get('investment_ratio', 10) / 100
            buy_amount = krw_balance * ratio

            if buy_amount < 5000:
                self.log_auto_trade(f"ℹ️ {ticker} 매수 건너뜀 (주문 금액 부족: {buy_amount:,.0f} KRW)")
                return

            price = pyupbit.get_current_price(ticker)
            self.log_auto_trade(f"📈 [{reason}] {ticker} 매수 신호 포착")
            result = upbit.buy_market_order(ticker, buy_amount)
            # 업비트 수수료율(시장가 0.05%) 적용
            fee = buy_amount * 0.0005
            # --- [추가] 매수 전략명을 저장 ---
            self.last_buy_strategy = {} if not hasattr(self, 'last_buy_strategy') else self.last_buy_strategy
            self.last_buy_strategy[ticker] = reason.split(":")[0].strip()  # 예: "전략1"
            self.log_auto_trade(
                f"✅ {ticker} {buy_amount:,.0f} KRW 만큼 매수 주문 완료. (주문결과: {result})",
                log_dict={
                    "datetime": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "ticker": ticker,
                    "side": "매수",
                    "reason": reason,
                    "entry_price": price,
                    "amount": round(buy_amount / price, 8) if price else "",
                    "total": buy_amount,
                    "exit_price": "",
                    "profit": "",
                    "profit_rate": "",
                    "fee": fee,
                    "profit_rate_net": ""
                }
            )
            self.last_sell_time[ticker] = datetime.now()
        except Exception as e:
            self.log_auto_trade(f"❗️ {ticker} 매수 주문 실행 중 오류: {e}")

    def check_sell_condition(self, ticker, coin_info):
        s = self.auto_trade_settings
        df = self.get_technical_indicators(ticker, interval='minute5', count=200)
        if df is None or len(df) < 10:
            return
        last = df.iloc[-1]
        prev = df.iloc[-2]

        # --- [변경] 매수 전략에 따라 매도 조건만 적용 ---
        strategy_map = {
            "전략1": lambda: last['rsi'] > 70 and last['ma5'] < last['ma20'],
            "전략2": lambda: (
                last['close'] > last['close'].rolling(20).mean().iloc[-1] + (last['close'].rolling(20).std().iloc[-1] * 2)
                and last['pattern'] == 'shooting_star'
            ),
            "전략3": lambda: (
                (prev['macd'] > prev['signal'] and last['macd'] < last['signal']) or
                (float(coin_info.get('avg_buy_price', 0)) > 0 and last['close'] < df['high'].iloc[-20:].max() * 0.97)
            ),
            "전략4": lambda: (not last['is_green'] and last['volume'] > prev['volume'] * 2),
            "전략5": lambda: (prev['ma5'] > prev['ma20'] and last['ma5'] < last['ma20']),
            "전략6": lambda: (
                'obv' in df.columns and prev['obv'] > last['obv']
            ),
            "전략7": lambda: (
                (last['rsi'] - last['rsi'].rolling(14, min_periods=1).min().iloc[-1]) /
                (last['rsi'].rolling(14, min_periods=1).max().iloc[-1] - last['rsi'].rolling(14, min_periods=1).min().iloc[-1]) <= 0.8
                if last['rsi'].rolling(14, min_periods=1).max().iloc[-1] != last['rsi'].rolling(14, min_periods=1).min().iloc[-1] else False
            ),
            "전략8": lambda: (
                ((last['high'] + last['low'] + last['close']) / 3 - ((last['high'] + last['low'] + last['close']) / 3).rolling(20, min_periods=1).mean().iloc[-1]) /
                (0.015 * ((last['high'] + last['low'] + last['close']) / 3).rolling(20, min_periods=1).apply(lambda x: np.mean(np.abs(x - np.mean(x)))).iloc[-1]) <= -100
                if ((last['high'] + last['low'] + last['close']) / 3).rolling(20, min_periods=1).apply(lambda x: np.mean(np.abs(x - np.mean(x)))).iloc[-1] != 0 else False
            ),
        }

        # 매수 시 사용한 전략만 매도 조건 적용
        buy_strategy = getattr(self, 'last_buy_strategy', {}).get(ticker)
        if not buy_strategy:
            return

        if buy_strategy in strategy_map and strategy_map[buy_strategy]():
            self.execute_sell(ticker, coin_info, f"{buy_strategy}: 매도조건")


    def create_buy_sell_tab(self, parent_frame, side):
        is_buy = (side == "buy")
        order_type_var = self.buy_order_type if is_buy else self.sell_order_type
        price_var = self.buy_price_var if is_buy else self.sell_price_var
        amount_var = self.buy_amount_var if is_buy else self.sell_amount_var
        total_var = self.buy_total_var if is_buy else self.sell_total_var
        balance_var = self.buy_krw_balance_var if is_buy else self.sell_coin_balance_var

        top_frame = ttk.Frame(parent_frame); top_frame.pack(fill='x', expand=True, pady=(0, 5))
        order_type_frame = ttk.Frame(top_frame); order_type_frame.pack(side='left')
        ttk.Radiobutton(order_type_frame, text="지정가", variable=order_type_var, value="limit", command=self._update_order_ui_state).pack(side="left")
        ttk.Radiobutton(order_type_frame, text="시장가", variable=order_type_var, value="market", command=self._update_order_ui_state).pack(side="left")
        ttk.Label(top_frame, textvariable=balance_var, foreground='gray').pack(side='right')

        grid_frame = ttk.Frame(parent_frame); grid_frame.pack(fill='x', expand=True); grid_frame.columnconfigure(1, weight=1)
        labels = ["주문가격(KRW)", "주문수량(COIN)", "주문총액(KRW)"]
        vars_entries = [
            (price_var, ttk.Entry(grid_frame, textvariable=price_var)),
            (amount_var, ttk.Entry(grid_frame, textvariable=amount_var)),
            (total_var, ttk.Entry(grid_frame, textvariable=total_var))
        ]
        for i, (label_text, (var, entry)) in enumerate(zip(labels, vars_entries)):
            ttk.Label(grid_frame, text=f"{label_text: <10}").grid(row=i, column=0, sticky='w', padx=5, pady=2)
            entry.grid(row=i, column=1, sticky='ew', padx=5, pady=2)
            if i==1:
                entry_symbol = ttk.Label(grid_frame, text="")
                entry_symbol.grid(row=i, column=2, sticky='w')
                if is_buy: self.buy_amount_symbol_label = entry_symbol
                else: self.sell_amount_symbol_label = entry_symbol

        entries = [e for _, e in vars_entries]
        if is_buy:
            self.buy_price_entry, self.buy_amount_entry, self.buy_total_entry = entries
        else:
            self.sell_price_entry, self.sell_amount_entry, self.sell_total_entry = entries

        percentage_frame = ttk.Frame(parent_frame)
        percentage_frame.pack(fill='x', expand=True, pady=5)
        if is_buy:
            for p in [0.1, 0.25, 0.5, 1.0]:
                btn = ttk.Button(percentage_frame, text=f"{p:.0%}", command=lambda pct=p: self._apply_buy_percentage(pct))
                btn.pack(side='left', fill='x', expand=True, padx=2)
        else:
            ttk.Label(percentage_frame, text="매도비율: ").pack(side="left")
            percentages = [f'{i}%' for i in range(5, 101, 5)]
            sell_combo = ttk.Combobox(percentage_frame, textvariable=self.sell_percentage_var, values=percentages, width=10)
            sell_combo.pack(side="left", padx=5)
            sell_combo.bind("<<ComboboxSelected>>", self._on_sell_percentage_select)

        action_text = "매수" if is_buy else "매도"
        style = "Buy.TButton" if is_buy else "Sell.TButton"
        ttk.Style().configure(style, foreground="black", background="#d24f45" if is_buy else "#1e6bde", font=('Helvetica', 10, 'bold'))
        action_button = ttk.Button(parent_frame, text=action_text, style=style, command=lambda s=side: self.place_order(s))
        action_button.pack(fill='x', expand=True, ipady=5, pady=(5,0))
        
    def create_history_tab(self, parent_frame):
        refresh_button = ttk.Button(parent_frame, text="거래내역 새로고침", command=self.on_refresh_history_click)
        refresh_button.pack(pady=(0, 10), fill='x')

        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(fill='both', expand=True)

        cols = ('datetime', 'name', 'side', 'price', 'qty', 'total')
        self.history_tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        col_map = {"거래시각": 140, "종목명": 100, "구분": 50, "체결가": 90, "수량": 100, "거래금액": 100}
        
        for i, (text, width) in enumerate(col_map.items()):
            self.history_tree.heading(cols[i], text=text)
            self.history_tree.column(cols[i], width=width, anchor='e')
        self.history_tree.column('name', anchor='w')
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.history_tree.pack(side="left", fill="both", expand=True)
        
    def load_trade_history_from_file(self):
        try:
            with open("trade_history.json", "r", encoding="utf-8") as f:
                self.trade_history_data = json.load(f)
                print(f"✅ 로컬 파일에서 거래내역 {len(self.trade_history_data)}건을 불러왔습니다.")
        except (FileNotFoundError, json.JSONDecodeError):
            self.trade_history_data = []
            print("ℹ️ 저장된 로컬 거래내역 파일이 없습니다.")

    def save_trade_history_to_file(self):
        with open("trade_history.json", "w", encoding="utf-8") as f:
            json.dump(self.trade_history_data, f, ensure_ascii=False, indent=4)

    def update_history_tree_gui(self):
        self.history_tree.delete(*self.history_tree.get_children())
        sorted_history = sorted(self.trade_history_data, key=lambda x: x['datetime'], reverse=True)
        for trade in sorted_history:
            self.history_tree.insert('', 'end', values=(
                trade['datetime'],
                trade['name'],
                trade['side'],
                f"{trade['price']:,.2f}",
                f"{trade['qty']:.8f}".rstrip('0').rstrip('.'),
                f"{trade['total']:,.0f}"
            ))

    def on_refresh_history_click(self):
        if messagebox.askyesno("거래내역 새로고침 확인", "기존에 저장된 내역에 추가로 최신 거래내역을 동기화합니다.\n종목 수가 많을 경우 시간이 오래 걸릴 수 있습니다.\n\n계속하시겠습니까?"):
            threading.Thread(target=self._fetch_and_update_trade_history_worker, daemon=True).start()

    def _fetch_and_update_trade_history_worker(self):
        self.after(0, lambda: messagebox.showinfo("조회 시작", "최신 거래 내역 동기화를 시작합니다. 완료되면 알려드립니다."))
        
        try:
            existing_uuids = {trade['uuid'] for trade in self.trade_history_data if 'uuid' in trade}
            new_trades = []
            
            tickers = pyupbit.get_tickers(fiat="KRW")
            print(f"🔍 총 {len(tickers)}개 마켓의 거래 내역 동기화를 시작합니다...")

            for i, ticker in enumerate(tickers):
                if (i + 1) % 10 == 0 or i == len(tickers) - 1:
                    print(f"  ({i+1}/{len(tickers)}) {ticker} 조회 중...")

                page = 1
                while True:
                    try:
                        orders = upbit.get_order(ticker, state="done", page=page, limit=100)
                        if not orders or not isinstance(orders, list):
                            break
                    except Exception as api_err:
                        print(f"❗️ {ticker} 거래내역 조회 API 오류: {api_err}")
                        break

                    for order in orders:
                        if order.get('uuid') in existing_uuids:
                            continue
                        
                        if 'trades' in order and order['trades']:
                            for trade in order['trades']:
                                new_trades.append({
                                    'uuid': order['uuid'],
                                    'datetime': pd.to_datetime(trade['created_at']).strftime('%Y-%m-%d %H:%M:%S'),
                                    'market': trade['market'],
                                    'name': self.ticker_to_display_name.get(trade['market'], trade['market']),
                                    'side': "매수" if trade['side'] == 'bid' else "매도",
                                    'price': float(trade['price']),
                                    'qty': float(trade['volume']),
                                    'total': float(trade['price']) * float(trade['volume'])
                                })
                    
                    page += 1
                    time.sleep(0.2)
            
            if not new_trades:
                self.after(0, lambda: messagebox.showinfo("동기화 완료", "새로운 거래 내역이 없습니다."))
                return

            self.trade_history_data.extend(new_trades)
            unique_trades = {d['uuid']: d for d in self.trade_history_data}
            self.trade_history_data = sorted(list(unique_trades.values()), key=lambda x: x['datetime'], reverse=True)

            self.save_trade_history_to_file()
            self.after(0, self.update_history_tree_gui)
            
            print(f"✅ 신규 거래 내역 {len(new_trades)}건을 추가했습니다.")
            self.after(0, lambda: messagebox.showinfo("동기화 완료", f"신규 거래 내역 {len(new_trades)}건을 성공적으로 동기화했습니다."))
        
        except Exception as e:
            print(f"❗️ 거래 내역 전체 동기화 중 오류 발생: {e}")
            self.after(0, lambda: messagebox.showerror("오류", f"거래 내역 동기화 중 오류가 발생했습니다.\n\n{e}"))

    def process_queue(self):
        try:
            while not self.data_queue.empty():
                task_name, data = self.data_queue.get_nowait()
                if task_name == "update_portfolio":
                    self.update_portfolio_gui(*data)
                elif task_name == "update_market":
                    self.market_data = data
                    self._refresh_market_tree_gui()
                elif task_name == "update_live_candle":
                    self._update_live_data(data)
                elif task_name == "draw_chart":
                    self._finalize_chart_drawing(*data)
                elif task_name == "draw_older_chart":
                    self._update_chart_after_loading(*data)
        except Empty:
            pass
        finally:
            if self.is_running:
                self.after(100, self.process_queue)
                
    def update_loop(self):
        if not self.is_running:
            return

        threading.Thread(target=self.fetch_current_price, daemon=True).start()

        if self.update_loop_counter % 5 == 0:
            threading.Thread(target=self._fetch_portfolio_data_worker, daemon=True).start()

        if self.update_loop_counter % 10 == 0:
            threading.Thread(target=self._fetch_market_data_worker, daemon=True).start()

        self.update_loop_counter += 1
        self.after(1000, self.update_loop)

    def fetch_current_price(self):
        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name)
        if ticker and ticker != "종목 없음":
            try:
                price = pyupbit.get_current_price(ticker)
                if price is not None:
                    self.data_queue.put(("update_live_candle", price))
            except Exception as e:
                print(f"❗️ 현재가 조회 오류: {e}")
                
    def _fetch_portfolio_data_worker(self):
        try:
            balances = upbit.get_balances()
            self.balances_data = {f"KRW-{b['currency']}": b for b in balances if b['currency'] != 'KRW'}
            krw_balance = next((float(b['balance']) for b in balances if b['currency'] == 'KRW'), 0.0)
            krw_balances_data = {f"KRW-{b['currency']}": b for b in balances if b['currency'] != 'KRW' and float(b.get('balance', 0)) > 0}

            display_name = self.selected_ticker_display.get()
            ticker = self.display_name_to_ticker.get(display_name)
            
            tickers_to_fetch = set(krw_balances_data.keys())
            if ticker:
                tickers_to_fetch.add(ticker)
            
            current_prices_dict = {}
            if tickers_to_fetch:
                price_data = pyupbit.get_current_price(list(tickers_to_fetch))
                if price_data:
                    current_prices_dict = price_data if isinstance(price_data, dict) else {list(tickers_to_fetch)[0]: price_data}
            
            total_investment, total_valuation = 0.0, 0.0
            portfolio_data_list = []

            for t, balance_info in krw_balances_data.items():
                balance = float(balance_info['balance'])
                avg_price = float(balance_info['avg_buy_price'])
                cur_price = current_prices_dict.get(t, 0)
                investment = balance * avg_price
                valuation = balance * cur_price
                total_investment += investment
                total_valuation += valuation
                portfolio_data_list.append({
                    'ticker': t, 'balance': balance, 'avg_price': avg_price,
                    'cur_price': cur_price, 'valuation': valuation, 'pl': valuation - investment
                })
            
            coin_balance = float(krw_balances_data.get(ticker, {}).get('balance', 0.0))
            coin_symbol = ticker.split('-')[1] if ticker and '-' in ticker else "COIN"

            total_pl = total_valuation - total_investment
            total_pl_rate = (total_pl / total_investment) * 100 if total_investment > 0 else 0
            
            result_data = (total_investment, total_valuation, total_pl, total_pl_rate, portfolio_data_list, krw_balance, coin_balance, coin_symbol)
            self.data_queue.put(("update_portfolio", result_data))
            
        except Exception as e:
            print(f"❗️ 포트폴리오 업데이트 오류: {e}")

    def _fetch_market_data_worker(self):
        try:
            all_tickers_krw = pyupbit.get_tickers(fiat="KRW")
            url = f"https://api.upbit.com/v1/ticker?markets={','.join(all_tickers_krw)}"
            response = requests.get(url)
            response.raise_for_status()
            market_data = response.json()
            if market_data:
                self.data_queue.put(("update_market", market_data))
        except Exception as e:
            print(f"❗️ KRW 마켓 목록 업데이트 중 오류: {e}")
    
    def on_ticker_select(self, event=None):
        self.draw_base_chart()
        self._update_order_ui_state()
        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name)
        if ticker:
            symbol = ticker.split('-')[1]
            self.buy_amount_symbol_label.config(text=symbol)
            self.sell_amount_symbol_label.config(text=symbol)
            
    def draw_base_chart(self, *args):
        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name, display_name)
        interval = self.selected_interval.get()
        if not ticker or ticker == "종목 없음": return
        # 이미 선택된 종목이면 확대/축소 상태 유지, 아니면 리셋
        if hasattr(self, "_keep_view") and self._keep_view:
            pass  # 확대/축소 상태 유지
        else:
            self._keep_view = False  # 다음 차트는 리셋
        self.master_df = None
        threading.Thread(target=self._fetch_and_draw_chart, args=(ticker, interval, display_name), daemon=True).start()

    def _fetch_and_draw_chart(self, ticker, interval, display_name):
        try:
            df = self.get_technical_indicators(ticker, interval=interval, count=200)
            self.data_queue.put(("draw_chart", (df, interval, display_name)))
        except Exception as e:
            print(f"❗️ 차트 데이터 로딩 오류: {e}")

    def _update_live_data(self, price):
        if self.master_df is None or self.master_df.empty:
            return
        
        self.current_price = price
        
        last_idx = self.master_df.index[-1]
        self.master_df.loc[last_idx, 'close'] = price
        if price > self.master_df.loc[last_idx, 'high']:
            self.master_df.loc[last_idx, 'high'] = price
        if price < self.master_df.loc[last_idx, 'low']:
            self.master_df.loc[last_idx, 'low'] = price

        self.update_overlays() # 오버레이만 업데이트
        self.canvas.draw_idle()

    def _finalize_chart_drawing(self, df, interval, display_name):
        self.master_df = df

        if self.master_df is None or len(self.master_df) < 2:
            self.ax.clear()
            self.ax.text(0.5, 0.5, "차트 데이터가 없습니다.", horizontalalignment='center', verticalalignment='center')
            self.canvas.draw()
            self.master_df = None
            self._keep_view = False
            return

        # 현재 확대/축소 상태 기억
        cur_xlim = self.ax.get_xlim()
        cur_ylim = self.ax.get_ylim()

        self._redraw_chart()

        # 확대/축소 상태가 유효하면 복원, 아니면 기본값
        if hasattr(self, "_keep_view") and self._keep_view and all(cur_xlim) and all(cur_ylim):
            try:
                self.ax.set_xlim(cur_xlim)
                self.ax.set_ylim(cur_ylim)
            except Exception:
                self.reset_chart_view()
        else:
            self.reset_chart_view()

        self.canvas.draw()
        self._keep_view = False  # 차트 그린 후에는 항상 False로 리셋

    # <<<<< [핵심 수정] 뷰 초기화 문제 해결을 위한 로직 재구성 >>>>>
    def _redraw_chart(self):
        self.ax.clear()
        for key in self.chart_elements: self.chart_elements[key].clear()
        
        if self.master_df is None or self.master_df.empty:
            self.canvas.draw()
            return
            
        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name, display_name)
        
        ma_data_to_plot, bb_data_to_plot = {}, {}
        
        for period, var in self.ma_vars.items():
            if var.get() and f'ma{period}' in self.master_df.columns:
                ma_data_to_plot[period] = self.master_df[f'ma{period}']
        
        if self.bb_var.get():
            bb_period = 20
            middle = self.master_df['close'].rolling(window=bb_period).mean()
            std = self.master_df['close'].rolling(window=bb_period).std()
            bb_data_to_plot = {'upper': middle + (std * 2), 'middle': middle, 'lower': middle - (std * 2)}
        
        trade_plots = self.prepare_trade_history_plots(ticker, self.master_df)
        
        current_interval = self.selected_interval.get()
        dt_format = '%m-%d %H:%M' if current_interval not in ['day', 'week'] else '%Y-%m-%d'
        
        mpf.plot(self.master_df, type='candle', ax=self.ax, style='yahoo',
                 ylabel='Price (KRW)', addplot=trade_plots, datetime_format=dt_format, xrotation=20)

        all_lows, all_highs = self.master_df['low'], self.master_df['high']

        data_min, data_max = all_lows.min(), all_highs.max()
        padding = (data_max - data_min) * 0.1
        y_bound_min = max(0, data_min - padding)
        y_bound_max = data_max + padding
        self.data_bounds = {'x': (0, len(self.master_df) - 1), 'y': (y_bound_min, y_bound_max)}
        
        self.plot_moving_averages(ma_data_to_plot)
        self.plot_bollinger_bands(bb_data_to_plot)
        
        self.ax.grid(True, linestyle='--', alpha=0.6)
        if trade_plots or ma_data_to_plot or bb_data_to_plot:
            self.ax.legend()
            
        self.update_overlays()
        print(f"📈 {display_name} ({self.selected_interval.get()}) 차트를 새로 그렸습니다. (총 {len(self.master_df)}개 캔들)")
        
    def _update_chart_after_loading(self, new_df, new_xlim):
        num_added = len(new_df) - (len(self.master_df) if self.master_df is not None else 0)
        self.master_df = new_df
        print(f"✅ 과거 캔들({num_added})을 추가했습니다. 총 {len(new_df)}개")
        self._redraw_chart()
        self.ax.set_xlim(new_xlim)
        self.canvas.draw()
        self.is_loading_older = False

    def is_cooldown(self, ticker, strategy_name, cooldown_minutes):
        """매수 쿨다운 중이면 True, 아니면 False. 쿨다운 남은 시간 로그 출력"""
        last_time = self.last_sell_time.get(ticker)
        if last_time:
            elapsed = datetime.now() - last_time
            remain = timedelta(minutes=cooldown_minutes) - elapsed
            if remain.total_seconds() > 0:
                mins, secs = divmod(int(remain.total_seconds()), 60)
                self.log_auto_trade(f"⏳ {ticker} {strategy_name} 매수조건 쿨다운 중 (남은시간: {mins:02d}:{secs:02d})")
                return True
        return False

    def _on_buy_input_change(self, *args):
        if self._is_calculating or self.buy_order_type.get() != 'limit': return
        self._is_calculating = True
        try:
            price = float(self.buy_price_var.get() or 0)
            amount = float(self.buy_amount_var.get() or 0)
            self.buy_total_var.set(f"{price * amount:.0f}")
        except (ValueError, TclError): pass
        finally: self._is_calculating = False

    def _on_buy_total_change(self, *args):
        if self._is_calculating or self.buy_order_type.get() != 'limit': return
        self._is_calculating = True
        try:
            price = float(self.buy_price_var.get() or 0)
            total = float(self.buy_total_var.get() or 0)
            if price > 0: self.buy_amount_var.set(f"{total / price:g}")
        except (ValueError, TclError): pass
        finally: self._is_calculating = False

    def _on_sell_input_change(self, *args):
        if self._is_calculating or self.sell_order_type.get() != 'limit': return
        self._is_calculating = True
        try:
            price = float(self.sell_price_var.get() or 0)
            amount = float(self.sell_amount_var.get() or 0)
            self.sell_total_var.set(f"{price * amount:.0f}")
        except (ValueError, TclError): pass
        finally: self._is_calculating = False

    def _update_order_ui_state(self):
        buy_type = self.buy_order_type.get()
        self.buy_price_entry.config(state='normal' if buy_type == 'limit' else 'disabled')
        self.buy_amount_entry.config(state='normal' if buy_type == 'limit' else 'disabled')
        self.buy_total_entry.config(state='normal' if buy_type == 'market' else 'disabled')
        if buy_type == 'market':
            self.buy_price_var.set("")
            self.buy_amount_var.set("")
        else: # limit
            self.buy_total_entry.config(state='normal')

        sell_type = self.sell_order_type.get()
        self.sell_price_entry.config(state='normal' if sell_type == 'limit' else 'disabled')
        self.sell_amount_entry.config(state='normal')
        self.sell_total_entry.config(state='disabled')
        if sell_type == 'market':
            self.sell_price_var.set("")
            self.sell_total_var.set("")

    def _apply_buy_percentage(self, percentage):
        buy_type = self.buy_order_type.get()
        total_krw = self.krw_balance * percentage
        if buy_type == 'limit':
            try:
                price = float(self.buy_price_var.get())
                if price > 0:
                    self.buy_amount_var.set(f"{total_krw / price * 0.9995:g}") # 수수료 고려
                else:
                    messagebox.showwarning("가격 입력 필요", "주문 가격을 먼저 입력해주세요.")
            except (ValueError, TclError):
                messagebox.showwarning("가격 입력 필요", "유효한 주문 가격을 먼저 입력해주세요.")
        else: # market
            self.buy_total_var.set(f"{total_krw:.0f}")

    def _on_sell_percentage_select(self, event=None):
        try:
            percentage_str = self.sell_percentage_var.get()
            if not percentage_str: return
            percentage = float(percentage_str.replace('%', '')) / 100
            amount = self.coin_balance * percentage
            self.sell_amount_var.set(f"{amount:g}")
            self.sell_percentage_var.set('')
        except (ValueError, TclError): pass

    def place_order(self, side):
        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name, display_name)
        is_buy = (side == "buy")
        order_type = self.buy_order_type.get() if is_buy else self.sell_order_type.get()

        try:
            if order_type == 'limit':
                price_str = self.buy_price_var.get() if is_buy else self.sell_price_var.get()
                amount_str = self.buy_amount_var.get() if is_buy else self.sell_amount_var.get()
                price = float(price_str)
                amount = float(amount_str)
            else: # market
                price = None
                if is_buy:
                    amount_str = self.buy_total_var.get()
                    amount = float(amount_str)
                else:
                    amount_str = self.sell_amount_var.get()
                    amount = float(amount_str)

            if amount <= 0:
                raise ValueError("주문수량/총액은 0보다 커야 합니다.")

        except (ValueError, TclError):
            messagebox.showerror("입력 오류", "유효한 숫자를 입력하세요.")
            return

        order_side_text = "매수" if is_buy else "매도"
        order_type_text = "지정가" if order_type == 'limit' else "시장가"
        price_text = f"주문 가격: {price:,.0f} KRW\n" if price is not None else ""
        
        if order_type == 'market' and is_buy:
            amount_label, amount_unit, amount_display = "주문 총액", "KRW", f"{amount:,.0f}"
        else:
            amount_label, amount_unit, amount_display = "주문 수량", ticker.split('-')[1], f"{amount:g}"

        confirm_msg = (f"[[ 주문 확인 ]]\n\n종목: {display_name}\n종류: {order_side_text} / {order_type_text}\n{price_text}{amount_label}: {amount_display} {amount_unit}\n\n위 내용으로 주문하시겠습니까?")

        if not messagebox.askyesno("주문 확인", confirm_msg):
            return
            
        try:
            result = None
            if is_buy:
                if order_type == 'limit': result = upbit.buy_limit_order(ticker, price, amount)
                else: result = upbit.buy_market_order(ticker, buy_amount)
            else:
                if order_type == 'limit': result = upbit.sell_limit_order(ticker, price, amount)
                else: result = upbit.sell_market_order(ticker, amount)

            messagebox.showinfo("주문 성공", f"주문이 성공적으로 접수되었습니다.\n\n{result}")
            self.buy_price_var.set(""); self.buy_amount_var.set(""); self.buy_total_var.set("")
            self.sell_price_var.set(""); self.sell_amount_var.set(""); self.sell_total_var.set("")

            threading.Thread(target=self._fetch_portfolio_data_worker, daemon=True).start()
            self.after(2000, self.on_refresh_history_click)

        except Exception as e:
            messagebox.showerror("주문 실패", f"주문 중 오류가 발생했습니다.\n\n{e}")
    
    def on_tree_double_click(self, event):
        item_id = self.portfolio_tree.focus()
        if not item_id: return
        item_values = self.portfolio_tree.item(item_id, "values")
        display_name = item_values[0]
        if not display_name: return
        # 이미 선택된 종목이면 확대/축소 상태 유지, 차트 재호출 방지
        # display_name이 마켓 트리뷰에 있으면 해당 row를 선택, 없으면 selection 해제
        found = False
        for iid in self.market_tree.get_children():
            vals = self.market_tree.item(iid, "values")
            if vals and vals[0] == display_name:
                self.market_tree.selection_set(iid)
                self.market_tree.focus(iid)
                found = True
                break
        if not found:
            self.market_tree.selection_remove(self.market_tree.selection())
            self.market_tree.focus('')

        self._ignore_market_select_event = False

        # [추가] 보유코인 트리뷰 selection도 명확히 표시
        for iid in self.portfolio_tree.get_children():
            vals = self.portfolio_tree.item(iid, "values")
            if vals and vals[0] == display_name:
                self.portfolio_tree.selection_set(iid)
                self.portfolio_tree.focus(iid)
                break

        print(f"📋 포트폴리오 더블클릭: {display_name} 차트를 표시합니다.")

    def on_market_list_select(self, event):
        # [핵심] 내부에서 programmatically 변경된 경우 이벤트 무시
        if getattr(self, '_ignore_market_select_event', False):
            return
        selection = self.market_tree.selection()
        if not selection: return
        item = self.market_tree.item(selection[0])
        display_name = item['values'][0]
        if self.selected_ticker_display.get() == display_name:
            self._keep_view = True
            return
        else:
            self._keep_view = False

        # [핵심] 보유코인 트리뷰 selection 해제 (동기화)
        self.portfolio_tree.selection_remove(self.portfolio_tree.selection())
        self.portfolio_tree.focus('')

        self.selected_ticker_display.set(display_name)
        self.on_ticker_select()
        print(f"💹 거래대금 목록 선택: {display_name} 차트를 표시합니다.")

    def reset_chart_view(self):
        if self.master_df is None or len(self.master_df) < 1: return
        
        view_start = max(0, len(self.master_df) - 200)
        view_end = len(self.master_df) + 2 
        self.ax.set_xlim(view_start, view_end - 1)

        try:
            visible_df = self.master_df.iloc[int(view_start):int(view_end-2)] 
            min_low = visible_df['low'].min()
            max_high = visible_df['high'].max()
            padding = (max_high - min_low) * 0.05 
            self.ax.set_ylim(min_low - padding, max_high + padding)
        except Exception as e:
            print(f"❗️ 뷰 리셋 중 Y축 범위 설정 오류: {e}")
            self.ax.autoscale(enable=True, axis='y', tight=False)

        self.canvas.draw_idle()
        print("🔄️ 차트 뷰를 초기 상태로 리셋했습니다.")

    def on_scroll(self, event):
        if event.inaxes != self.ax: return
        zoom_factor = 1 / 1.1 if event.step > 0 else 1.1
        cur_xlim = self.ax.get_xlim(); cur_ylim = self.ax.get_ylim()
        x_data, y_data = event.xdata, event.ydata
        if x_data is None or y_data is None: return
        new_xlim = [(cur_xlim[0] - x_data) * zoom_factor + x_data, (cur_xlim[1] - x_data) * zoom_factor + x_data]
        new_ylim = [(cur_ylim[0] - y_data) * zoom_factor + y_data, (cur_ylim[1] - y_data) * zoom_factor + y_data]
        x_bounds = self.data_bounds.get('x'); y_bounds = self.data_bounds.get('y')
        if x_bounds:
            if new_xlim[0] < x_bounds[0]: new_xlim[0] = x_bounds[0]
            if new_xlim[1] > x_bounds[1] + 2: new_xlim[1] = x_bounds[1] + 2 
        if y_bounds:
            if new_ylim[0] < y_bounds[0]: new_ylim[0] = y_bounds[0]
            if new_ylim[1] > y_bounds[1]: new_ylim[1] = y_bounds[1]
        self.ax.set_xlim(new_xlim); self.ax.set_ylim(new_ylim); self.canvas.draw_idle()
    
    def on_press(self, event):
        if event.inaxes != self.ax: return
        if event.dblclick:
            self.reset_chart_view()
            return
        self.is_panning = True
        self.pan_start_pos = (event.xdata, event.ydata)

    def on_motion(self, event):
        if not self.is_panning or event.inaxes != self.ax or self.pan_start_pos is None: return
        try:
            dx = event.xdata - self.pan_start_pos[0]
            dy = event.ydata - self.pan_start_pos[1]
            cur_xlim = self.ax.get_xlim(); cur_ylim = self.ax.get_ylim()
            new_xlim = [cur_xlim[0] - dx, cur_xlim[1] - dx]
            x_bounds = self.data_bounds.get('x')
            if x_bounds:
                width = new_xlim[1] - new_xlim[0]
                if new_xlim[0] < x_bounds[0]: new_xlim = [x_bounds[0], x_bounds[0] + width]
                if new_xlim[1] > x_bounds[1] + 2: new_xlim = [x_bounds[1] + 2 - width, x_bounds[1] + 2]
            self.ax.set_xlim(new_xlim); self.ax.set_ylim([cur_ylim[0] - dy, cur_ylim[1] - dy]); self.canvas.draw_idle()
        except TypeError: pass

    def on_release(self, event):
        if not self.is_panning: return
        self.is_panning = False; self.pan_start_pos = None
        current_xlim = self.ax.get_xlim()
        if current_xlim[0] < 1 and not self.is_loading_older:
            print("⏳ 차트 왼쪽 끝에 도달, 과거 데이터를 로딩합니다...");
            self.load_older_data()

    def load_older_data(self):
        if self.master_df is None or self.master_df.empty: return
        self.is_loading_older = True
        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name)
        interval = self.selected_interval.get()
        to_date = self.master_df.index[0]
        current_xlim = self.ax.get_xlim()
        threading.Thread(target=self._fetch_older_data_worker, args=(ticker, interval, to_date, current_xlim), daemon=True).start()

    # <<<<< [핵심 수정] 과거 데이터 로딩 오류 해결 >>>>>
    def _fetch_older_data_worker(self, ticker, interval, to_date, current_xlim):
            try:
                # pyupbit의 to 파라미터는 해당 시점 '이전'까지 200개를 반환하므로, to_date에서 1초 빼기
                if isinstance(to_date, pd.Timestamp):
                    to_date_str = (to_date - pd.Timedelta(seconds=1)).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    to_date_str = pd.to_datetime(to_date) - pd.Timedelta(seconds=1)
                    to_date_str = to_date_str.strftime('%Y-%m-%d %H:%M:%S')

                # 과거 데이터 요청 (최대 200개)
                older_df_raw = pyupbit.get_ohlcv(ticker, interval=interval, count=200, to=to_date_str)
                # 만약 데이터가 없거나 1개 이하만 반환되면, 더 이상 로드할 데이터가 없음
                if older_df_raw is None or len(older_df_raw) < 2:
                    print("ℹ️ 더 이상 로드할 과거 데이터가 없습니다.")
                    self.is_loading_older = False
                    return

                # 현재 데이터에서 순수 OHLCV만 추출
                current_ohlcv = self.master_df[['open', 'high', 'low', 'close', 'volume']]
                # 중복 인덱스 제거 (과거 데이터와 현재 데이터가 겹칠 수 있음)
                combined_df_raw = pd.concat([older_df_raw, current_ohlcv])
                combined_df_raw = combined_df_raw[~combined_df_raw.index.duplicated(keep='last')]
                # 인덱스 정렬
                combined_df_raw = combined_df_raw.sort_index()

                # 지표 계산 최소 데이터 개수 완화 (신규상장 코인 대응)
                df_with_indicators = self.get_technical_indicators_from_raw(combined_df_raw, min_length=2)
                # 기존 master_df보다 더 많은 데이터가 있으면 갱신
                if df_with_indicators is not None and not df_with_indicators.empty:
                    num_candles_added = len(df_with_indicators) - len(self.master_df)
                    if num_candles_added > 0:
                        # xlim을 왼쪽으로 이동
                        new_xlim = (current_xlim[0] + num_candles_added, current_xlim[1] + num_candles_added)
                        self.data_queue.put(("draw_older_chart", (df_with_indicators, new_xlim)))
                        return
                print("ℹ️ 더 이상 로드할 과거 데이터가 없습니다.")
            except Exception as e:
                print(f"❗️ 과거 데이터 로딩 중 오류 발생: {e}")
            self.is_loading_older = False
    
    def prepare_trade_history_plots(self, ticker, chart_df):
        buy_signals, sell_signals = [], []
        try:
            if not self.trade_history_data: return []
            buy_prices = pd.Series([np.nan] * len(chart_df), index=chart_df.index)
            sell_prices = pd.Series([np.nan] * len(chart_df), index=chart_df.index)
            ticker_history = [t for t in self.trade_history_data if t['market'] == ticker]
            for trade in ticker_history:
                trade_time = pd.to_datetime(trade['datetime'])

                closest_time = chart_df.index.asof(trade_time)
                if pd.notnull(closest_time):
                    loc = chart_df.index.get_loc(closest_time)
                    if trade['side'] == '매수':
                        buy_prices.iloc[loc] = chart_df['high'].iloc[loc] * 1.02
                    else:
                        sell_prices.iloc[loc] = chart_df['low'].iloc[loc] * 0.98
            if not buy_prices.dropna().empty:
                buy_signals.append(mpf.make_addplot(buy_prices, type='scatter', marker='^', color='red', markersize=80, label='매수(Buy)'))
            if not sell_prices.dropna().empty:
                sell_signals.append(mpf.make_addplot(sell_prices, type='scatter', marker='v', color='blue', markersize=80, label='매도(Sell)'))
            return buy_signals + sell_signals
        except Exception as e:
            print(f"❗️ 매매 내역 표시 중 오류: {e}")
            return []
            
    def plot_moving_averages(self, ma_data):
        if ma_data:
            lines = []
            for period, ma_series in ma_data.items():
                line, = self.ax.plot(range(len(ma_series)), ma_series.values, label=f"MA{period}")
                lines.append(line)
            self.chart_elements['main'].extend(lines)

    def plot_bollinger_bands(self, bb_data):
        if bb_data:
            x_axis = range(len(self.master_df))
            middle, = self.ax.plot(x_axis, bb_data['middle'].values, color='orange', linestyle='--', linewidth=1, label='BB Center')
            upper, = self.ax.plot(x_axis, bb_data['upper'].values, color='gray', linestyle='--', linewidth=0.7)
            lower, = self.ax.plot(x_axis, bb_data['lower'].values, color='gray', linestyle='--', linewidth=0.7)
            fill = self.ax.fill_between(x_axis, bb_data['lower'].values, bb_data['upper'].values, color='gray', alpha=0.1)
            self.chart_elements['main'].extend([middle, upper, lower, fill])
    
    def update_portfolio_gui(self, total_investment, total_valuation, total_pl, total_pl_rate, portfolio_data, krw_balance, coin_balance, coin_symbol):
        self.krw_balance = krw_balance
        self.coin_balance = coin_balance

        # --- [추가] 종합 현황 텍스트 갱신 ---
        self.krw_balance_summary_var.set(f"보유 KRW: {krw_balance:,.0f} 원")
        self.total_investment_var.set(f"총 투자금액: {total_investment:,.0f} 원")
        self.total_valuation_var.set(f"총 평가금액: {total_valuation:,.0f} 원")
        self.total_pl_var.set(f"총 평가손익: {total_pl:,.0f} 원 ({total_pl_rate:+.2f}%)")

        # --- [기존] 보유코인 트리뷰 갱신 ---
        self.portfolio_tree.delete(*self.portfolio_tree.get_children())
        for item in portfolio_data:
            display_name = self.ticker_to_display_name.get(item['ticker'], item['ticker'])
            balance = item['balance']
            avg_price = item['avg_price']
            cur_price = item['cur_price']
            valuation = item['valuation']
            pl = item['pl']
            pl_rate = (pl / (avg_price * balance) * 100) if avg_price > 0 and balance > 0 else 0
            tag = 'plus' if pl > 0 else 'minus' if pl < 0 else ''
            self.portfolio_tree.insert(
                '', 'end',
                values=(

                    display_name,
                    f"{balance:.8f}".rstrip('0').rstrip('.'),
                    f"{avg_price:,.2f}",
                    f"{cur_price:,.2f}",
                    f"{valuation:,.0f}",
                    f"{pl:,.0f}",
                    f"{pl_rate:+.2f}%"
                ),
                tags=(tag,)
            )
        # --- 이하 기존 코드(코인 비중, 잔고 등) 계속 ---
        if portfolio_data and total_valuation > 0:
            chart_data = [{'label': self.ticker_to_display_name.get(item['ticker'], item['ticker']), 'value': item['valuation']} for item in portfolio_data]
            main_items = []
            other_items_value = 0.0  # ← 리스트가 아니라 float로 초기화
            sorted_chart_data = sorted(chart_data, key=lambda x: x['value'], reverse=True)

            for item in sorted_chart_data:
                percentage = (item['value'] / total_valuation) * 100
                if len(main_items) < 7 and percentage >= 2.0:
                    main_items.append(item)
                else:
                    other_items_value += item['value']  # ← 리스트 append가 아니라 누적합

            if other_items_value > 0:
                main_items.append({'label': '기타', 'value': other_items_value})

            labels = [item['label'].split('(')[0] for item in main_items][::-1]
            percentages = [(item['value'] / total_valuation) * 100 for item in main_items][::-1]
            
            num_items = len(labels)
            try:
                colors = plt.colormaps.get_cmap('viridis_r')(np.linspace(0, 1, num_items))
            except AttributeError:
                colors = plt.cm.get_cmap('viridis_r', num_items)( range(num_items))
            
            bars = self.pie_ax.barh(labels, percentages, color=colors, height=0.6)
            self.pie_ax.set_xlabel('비중 (%)', fontsize=9)
            self.pie_ax.tick_params(axis='y', labelsize=9)
            self.pie_ax.set_title('포트폴리오 구성', fontsize=11, fontweight='bold')

            self.pie_ax.spines['top'].set_visible(False)
            self.pie_ax.spines['right'].set_visible(False)
            self.pie_ax.spines['left'].set_visible(False)
            
            for bar in bars:
                width = bar.get_width()
                self.pie_ax.text(width + 0.5, bar.get_y() + bar.get_height()/2.,
                                 f'{width:.1f}%', ha='left', va='center', fontsize=8.5)

            self.pie_ax.set_xlim(0, max(percentages) * 1.15 if percentages else 100)
        else:
            self.pie_ax.text(0.5, 0.5, "보유 코인이 없습니다", horizontalalignment='center', verticalalignment='center')
            self.pie_ax.set_xticks([]); self.pie_ax.set_yticks([])

        self.pie_fig.tight_layout()
        self.pie_canvas.draw()
        
        self.buy_krw_balance_var.set(f"주문가능: {krw_balance:,.0f} KRW")
        self.sell_coin_balance_var.set(f"주문가능: {coin_balance:g} {coin_symbol}")

    def _refresh_market_tree_gui(self):
        if not self.market_data: return
        sort_key_map = { 'display_name': 'market', 'price': 'trade_price', 'change_rate': 'signed_change_rate', 'volume': 'acc_trade_price_24h' }
        key_to_sort = sort_key_map.get(self.sort_column, 'acc_trade_price_24h')
        sorted_data = sorted(self.market_data, key=lambda x: x[key_to_sort], reverse=not self.sort_ascending)
        try:
            selected_id = self.market_tree.focus()
            selected_display_name = self.market_tree.item(selected_id, 'values')[0] if selected_id else None
            self.market_tree.delete(*self.market_tree.get_children()); new_selection_id = None
            for item in sorted_data:
                ticker_name = item['market']; display_name = self.ticker_to_display_name.get(ticker_name, ticker_name)
                price = item['trade_price']; change_rate = item['signed_change_rate'] * 100; volume = item['acc_trade_price_24h']
                tag = 'red' if change_rate > 0 else 'blue' if change_rate < 0 else 'black'
                price_str = f"{price:,.0f}" if price >= 100 else f"{price:g}"; change_rate_str = f"{change_rate:+.2f}%"; volume_str = self.format_trade_volume(volume)
                item_id = self.market_tree.insert('', 'end', values=(display_name, price_str, change_rate_str, volume_str), tags=(tag,))
                if display_name == selected_display_name: new_selection_id = item_id
            if new_selection_id:
                self.market_tree.focus(new_selection_id); self.market_tree.selection_set(new_selection_id)
        except Exception: pass

    def sort_market_list(self, col):
        if self.sort_column == col: self.sort_ascending = not self.sort_ascending
        else: self.sort_column = col; self.sort_ascending = False
        self._refresh_market_tree_gui()

    def format_trade_volume(self, volume):
        if volume > 1_000_000_000_000: return f"{volume / 1_000_000_000_000:.1f}조"
        if volume > 1_000_000_000: return f"{volume / 1_000_000_000:.0f}십억"
        if volume > 1_000_000: return f"{volume / 1_000_000:.0f}백만"
        return f"{volume:,.0f}"
    
       
    def load_my_tickers(self):
        threading.Thread(target=self._load_my_tickers_worker, daemon=True).start()

    def _load_my_tickers_worker(self):
        balances = upbit.get_balances()
        my_tickers = [f"KRW-{b['currency']}" for b in balances if b['currency'] != 'KRW' and float(b.get('balance', 0)) > 0]
        
        all_display_names = sorted(list(self.display_name_to_ticker.keys()))
        self.after(0, lambda: self.ticker_combobox.config(values=all_display_names))

        if my_tickers:
            first_ticker = my_tickers[0]
            display_name = self.ticker_to_display_name.get(first_ticker, first_ticker)
            self.selected_ticker_display.set(display_name)
        elif all_display_names:
            self.selected_ticker_display.set(all_display_names[0])
        else:
            self.selected_ticker_display.set("종목 없음")
        
        self.after(0, self.on_ticker_select)
        
    def update_overlays(self):
        for element in self.chart_elements['overlay']:
            try: element.remove()
            except: pass
        self.chart_elements['overlay'].clear()

        if self.master_df is None: return

        display_name = self.selected_ticker_display.get()
        ticker = self.display_name_to_ticker.get(display_name)
        if not ticker: return

        current_xlim = self.ax.get_xlim()
        current_ylim = self.ax.get_ylim()

        self.avg_buy_price = float(self.balances_data.get(ticker, {}).get('avg_buy_price', 0.0))
        profit_rate = ((self.current_price - self.avg_buy_price) / self.avg_buy_price) * 100 if self.avg_buy_price > 0 and self.current_price > 0 else 0
        self.ax.set_title(f'{display_name} ({self.selected_interval.get()}) Chart (수익률: {profit_rate:+.2f}%)', fontsize=14)
        
        right_limit = current_xlim[1]

        if self.current_price > 0 and current_ylim[0] < self.current_price < current_ylim[1]:
            line = self.ax.axhline(y=self.current_price, color='red', linestyle='--')
            text = self.ax.text(right_limit, self.current_price, f' {self.current_price:,.2f}', color='red', va='center', ha='left', bbox=dict(facecolor='white', alpha=0.5, edgecolor='none', boxstyle='round,pad=0.2'))
            self.chart_elements['overlay'].extend([line, text])
            
        if self.avg_buy_price > 0 and current_ylim[0] < self.avg_buy_price < current_ylim[1]:
            line = self.ax.axhline(y=self.avg_buy_price, color='blue', linestyle=':')
            text = self.ax.text(right_limit, self.avg_buy_price, f' {self.avg_buy_price:,.2f}', color='blue', va='center', ha='left', bbox=dict(facecolor='white', alpha=0.5, edgecolor='none', boxstyle='round,pad=0.2'))
            self.chart_elements['overlay'].extend([line, text])
        
    def on_closing(self):
        """프로그램 종료 시 호출되는 함수"""
        # 필요시, 종료 전 저장할 작업이 있으면 여기에 추가
        self.is_running = False
        if self.auto_trade_monitor is not None and self.auto_trade_monitor.winfo_exists():
            self.auto_trade_monitor.close()
        self.destroy()

class AutoTradeSettingsWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.master_app = master
        self.title("자동매매 시나리오 설정")
        self.geometry("700x600")
        self.resizable(False, False)

        self.vars = {
            'selected_ticker': tk.StringVar(),  # 드롭다운용 변수
            'investment_ratio': tk.StringVar(),
            'strategy1': tk.BooleanVar(),
            'strategy2': tk.BooleanVar(),
            'strategy3': tk.BooleanVar(),
            'strategy4': tk.BooleanVar(),
            'strategy5': tk.BooleanVar(),
            'strategy6': tk.BooleanVar(),
            'strategy7': tk.BooleanVar(),
            'strategy8': tk.BooleanVar(),
        }
        self.setup_widgets()
        self.load_settings()

        # [추가] 종목 선택 시 전략 자동 분석 및 체크
        self.vars['selected_ticker'].trace_add("write", self.on_ticker_selected)

    def on_ticker_selected(self, *args):
        display_name = self.vars['selected_ticker'].get()
        ticker = self.master_app.display_name_to_ticker.get(display_name)
        if not ticker:
            return

        try:
            df_raw = pyupbit.get_ohlcv(ticker, interval='minute5', count=200)
            if df_raw is None or len(df_raw) < 30:
                return

            # 기술적 지표 추가
            df = self.master_app.get_technical_indicators_from_raw(df_raw)
            if df is None or len(df) < 30:
                return

            close = df['close']
            ma60 = close.rolling(window=60).mean()
            ma20 = close.rolling(window=20).mean()
            ma5 = close.rolling(window=5).mean()
            last = df.iloc[-1]
            prev = df.iloc[-2]

            # 모든 전략 체크 해제
            for k in range(1, 9):
                self.vars[f'strategy{k}'].set(False)

            # 1. RSI+이동평균+거래량 (전략1) - 조건 완화
            vol_ma10 = df['volume'].rolling(10).mean().iloc[-1]
            if (last['rsi'] < 35 or last['ma5'] > last['ma20']) and last['volume'] > vol_ma10 * 1.5:
                self.vars['strategy1'].set(True)

            # 2. 볼린저밴드+캔들패턴 (전략2) - 조건 완화
            bb_period = 20
            middle = close.rolling(window=bb_period).mean()
            std = close.rolling(window=bb_period).std()
            lower = middle + (std * -2)
            if last['close'] < lower.iloc[-1] * 1.05 or last.get('pattern', '') == 'hammer':
                self.vars['strategy2'].set(True)

            # 3. MACD+트레일링스탑 (전략3) - 최근 3봉 중 1봉이라도 돌파
            macd_cross = any(
                df['macd'].iloc[-i] > df['signal'].iloc[-i] and df['macd'].iloc[-i-1] < df['signal'].iloc[-i-1]
                for i in range(1, 4)
            )
            if macd_cross:
                self.vars['strategy3'].set(True)

            # 4. 강한캔들+볼륨펌핑 (전략4) - 거래량 2배, 몸통 상위 30%
            body_threshold = df['body'].quantile(0.7)
            if last.get('is_green', False) and last['volume'] > prev['volume'] * 2 and last['body'] > body_threshold:
                self.vars['strategy4'].set(True)

            # 5. MA 골든/데드크로스 (전략5)
            if prev['ma5'] < prev['ma20'] and last['ma5'] > last['ma20']:
                self.execute_buy(ticker, "전략5: MA 골든크로스")
                return
            # 데드크로스 매도
            coin_info = self.master_app.balances_data.get(ticker)
            if coin_info and float(coin_info.get('balance', 0)) > 0:
                if prev['ma5'] > prev['ma20'] and last['ma5'] < last['ma20']:
                    self.execute_sell(ticker, coin_info, "전략5: MA 데드크로스")

            # --- [추가] 차트 형태 및 전략 로그 출력 ---
            # 1. 차트 형태 판별 (최근 20봉)
            close20 = df['close'].iloc[-20:]
            change_rate = (close20.iloc[-1] - close20.iloc[0]) / close20.iloc[0] * 100
            if change_rate > 5:
                chart_type = "상승 차트"
            elif change_rate < -5:
                chart_type = "하락 차트"
            else:
                chart_type = "횡보 차트"

            # 2. 선택된 전략 목록
            selected_strategies = []
            for k in range(1, 9):
                if self.vars[f'strategy{k}'].get():
                    selected_strategies.append(f"전략{k}")

            log_msg = f"{chart_type}입니다.\n차트 분석 결과 " + (", ".join(selected_strategies) + "가 선택되었습니다." if selected_strategies else "선택된 전략이 없습니다.")
            self.master_app.log_auto_trade(log_msg)

        except Exception as e:
            print(e)

    def setup_widgets(self):
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # [수정] 자동매매 대상 종목 드롭다운
        tickers_frame = ttk.LabelFrame(main_frame, text="[1] 자동매매 대상 종목 (1개만 선택)", padding=10)
        tickers_frame.pack(fill=tk.X, pady=5)

        # 보유/미보유 코인 모두 리스트업
        all_tickers = list(self.master_app.ticker_to_display_name.items())
        all_tickers.sort(key=lambda x: x[1])
        display_names = [v for k, v in all_tickers]
        self.ticker_combo = ttk.Combobox(tickers_frame, textvariable=self.vars['selected_ticker'], values=display_names, state="readonly", width=30)
        self.ticker_combo.pack(fill=tk.X, padx=5, pady=5)

        # 이하 기존 옵션(매수비중, 전략 등) 동일하게 배치
        common_frame = ttk.LabelFrame(main_frame, text="[2] 공통 옵션", padding=10)
        common_frame.pack(fill=tk.X, pady=5)
        ttk.Label(common_frame, text="1회 매수 비중:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        ratios = [f'{i}%' for i in range(10, 101, 10)]
        ratio_combo = ttk.Combobox(common_frame, textvariable=self.vars['investment_ratio'], values=ratios, width=10, state="readonly")
        ratio_combo.grid(row=0, column=1, sticky='w', padx=5)
        ttk.Label(common_frame, text="재매수 금지(쿨다운): 전략별 자동 적용").grid(row=0, column=2, columnspan=3, sticky='w', padx=5, pady=2)

        strategy_frame = ttk.LabelFrame(main_frame, text="[3] 단타 전략 선택 (하나 이상)", padding=10)
        strategy_frame.pack(fill=tk.X, pady=5)
        cb1 = ttk.Checkbutton(strategy_frame, text="전략 1: RSI+이동평균+거래량 (5분봉, 저위험)", variable=self.vars['strategy1'])
        cb1.pack(anchor='w', pady=2)
        cb2 = ttk.Checkbutton(strategy_frame, text="전략 2: 볼린저밴드+캔들패턴 (1분봉, 고수익)", variable=self.vars['strategy2'])
        cb2.pack(anchor='w', pady=2)
        cb3 = ttk.Checkbutton(strategy_frame, text="전략 3: MACD+트레일링스탑 (5분봉, 추세추종)", variable=self.vars['strategy3'])
        cb3.pack(anchor='w', pady=2)
        cb4 = ttk.Checkbutton(strategy_frame, text="전략 4: 강한캔들+볼륨펌핑 (1분봉, 초단타)", variable=self.vars['strategy4'])
        cb4.pack(anchor='w', pady=2)
        cb5 = ttk.Checkbutton(strategy_frame, text="전략 5: MA 골든/데드크로스 (5/20MA)", variable=self.vars['strategy5'])
        cb5.pack(anchor='w', pady=2)
        cb6 = ttk.Checkbutton(strategy_frame, text="전략 6: OBV 추세전환", variable=self.vars['strategy6'])
        cb6.pack(anchor='w', pady=2)
        cb7 = ttk.Checkbutton(strategy_frame, text="전략 7: StochRSI 돌파", variable=self.vars['strategy7'])
        cb7.pack(anchor='w', pady=2)
        cb8 = ttk.Checkbutton(strategy_frame, text="전략 8: CCI 돌파", variable=self.vars['strategy8'])
        cb8.pack(anchor='w', pady=2)

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side="bottom", pady=(5, 0))
        ttk.Button(button_frame, text="저장", command=self.save_and_close, style="On.TButton").pack(side="left", padx=10, ipady=4)
        ttk.Button(button_frame, text="닫기", command=self.destroy).pack(side="left", padx=10, ipady=4)

    def load_settings(self):
        s = self.master_app.auto_trade_settings
        # enabled_tickers가 있으면 첫 번째만 선택
        enabled = s.get('enabled_tickers', [])
        if enabled:
            ticker = enabled[0]
            display_name = self.master_app.ticker_to_display_name.get(ticker, ticker)
            self.vars['selected_ticker'].set(display_name)
        else:
            self.vars['selected_ticker'].set("")
        self.vars['investment_ratio'].set(f"{s.get('investment_ratio', 10)}%")
        self.vars['strategy1'].set(s.get('strategy1', False))
        self.vars['strategy2'].set(s.get('strategy2', False))
        self.vars['strategy3'].set(s.get('strategy3', False))
        self.vars['strategy4'].set(s.get('strategy4', False))
        self.vars['strategy5'].set(s.get('strategy5', False))
        self.vars['strategy6'].set(s.get('strategy6', False))
        self.vars['strategy7'].set(s.get('strategy7', False))
        self.vars['strategy8'].set(s.get('strategy8', False))

    def save_and_close(self):
        try:
            new = {}
            # 선택된 종목 1개만 enabled_tickers에 저장
            selected_display = self.vars['selected_ticker'].get()
            ticker = self.master_app.display_name_to_ticker.get(selected_display)
            new['enabled_tickers'] = [ticker] if ticker else []
            new['investment_ratio'] = int(self.vars['investment_ratio'].get().replace('%', ''))
            # 아래처럼 bool()로 감싸서 저장
            new['strategy1'] = bool(self.vars['strategy1'].get())
            new['strategy2'] = bool(self.vars['strategy2'].get())
            new['strategy3'] = bool(self.vars['strategy3'].get())
            new['strategy4'] = bool(self.vars['strategy4'].get())
            new['strategy5'] = bool(self.vars['strategy5'].get())
            new['strategy6'] = bool(self.vars['strategy6'].get())
            new['strategy7'] = bool(self.vars['strategy7'].get())
            new['strategy8'] = bool(self.vars['strategy8'].get())
            # 미보유 코인 전체 자동매수 옵션은 제거
            new['is_unowned_buy_enabled'] = False

            self.master_app.auto_trade_settings = new
            self.master_app.save_auto_trade_settings()
            messagebox.showinfo("저장 완료", "자동매매 설정이 저장되었습니다.", parent=self)
            self.destroy()
        except ValueError:
            messagebox.showerror("입력 오류", "입력된 값을 확인해주세요. 숫자 필드에는 숫자만 입력해야 합니다.", parent=self)
        except Exception as e:
            messagebox.showerror("오류", f"설정 저장 중 오류가 발생했습니다: {e}", parent=self)

class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        if self.tooltip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left',
                      background="#ffffe0", relief='solid', borderwidth=1,
                      font=("tahoma", "8", "normal"), wraplength=400)
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

    def show_tooltip(self, event):
        if self.tooltip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left',
                      background="#ffffe0", relief='solid', borderwidth=1,
                      font=("tahoma", "8", "normal"), wraplength=400)
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

    def show_tooltip(self, event):
        if self.tooltip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left',
                      background="#ffffe0", relief='solid', borderwidth=1,
                      font=("tahoma", "8", "normal"), wraplength=400)
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

class AutoTradeMonitorWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("자동매매 실시간 모니터링")
        self.geometry("1200x500")
        self.resizable(True, True)
        self.master_app = master

        columns = [
            "datetime", "ticker", "side", "reason", "entry_price", "amount", "total",
            "exit_price", "profit", "profit_rate", "fee", "profit_rate_net"
        ]
        headers = [
            "일시", "종목", "구분", "매수/매도 이유", "진입가", "수량", "매수금액",
            "청산가", "수익금액", "수익률(%)", "수수료", "수수료제외수익률(%)"
        ]
        self.columns = columns

        frame = ttk.Frame(self, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(frame, columns=columns, show="headings", height=18)
        for col, header in zip(columns, headers):
            self.tree.heading(col, text=header)
            self.tree.column(col, width=100, anchor='center')
        self.tree.pack(fill=tk.BOTH, expand=True, side="left")

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        self.log_data = []
        self.excel_filename = f"auto_trade_log_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self._init_excel()

    def _init_excel(self):
        try:
            # 파일이 없으면 새로 생성, 있으면 이어쓰기
            try:
                self.wb = openpyxl.load_workbook(self.excel_filename)
                self.ws = self.wb.active
            except FileNotFoundError:
                self.wb = openpyxl.Workbook()
                self.ws = self.wb.active
                self.ws.append([
                    "일시", "종목", "구분", "매수/매도 이유", "진입가", "수량", "매수금액",
                    "청산가", "수익금액", "수익률(%)", "수수료", "수수료제외수익률(%)"
                ])
                for i in range(1, 13):
                    self.ws.column_dimensions[get_column_letter(i)].width = 15
                self.wb.save(self.excel_filename)
        except Exception as e:
            print(f"❗️ 엑셀 파일 초기화 오류: {e}")

    def add_log(self, log_dict):
        # 트리뷰와 엑셀 모두에 기록
        def fmt(val, is_money=False):
            try:
                if val == "" or val is None:
                    return ""
                if is_money:
                    return f"{float(val):,.2f}"
                else:
                    return f"{float(val):.2f}"
            except Exception:
                return val

        # 소숫점 둘째자리로 표시할 컬럼명
        money_cols = {"entry_price", "total", "exit_price", "profit", "fee"}
        qty_cols = {"amount"}
        percent_cols = {"profit_rate", "profit_rate_net"}

        values = []
        for col in self.columns:
            v = log_dict.get(col, "")
            if col in money_cols:
                values.append(fmt(v, is_money=True))
            elif col in qty_cols:
                values.append(fmt(v))
            elif col in percent_cols:
                try:
                    values.append(f"{float(v):.2f}")
                except Exception:
                    values.append(v)
            else:
                values.append(v)

        self.tree.insert('', 0, values=values)
        self.log_data.append(log_dict)
        try:
            self.ws.append(values)
            self.wb.save(self.excel_filename)
        except Exception as e:
            print(f"❗️ 엑셀 저장 오류: {e}")

    def close(self):
        try:
            self.wb.save(self.excel_filename)
        except Exception:
            pass
        self.destroy()

if __name__ == "__main__":
    app = UpbitChartApp()
    app.start_updates()
    app.mainloop()